import os
import pickle
import math
import numpy as np
import pandas as pd
from sklearn.svm import SVC, SVR
from sklearn.metrics import accuracy_score, mean_squared_error
from sklearn.model_selection import train_test_split, KFold
from openpyxl import load_workbook, Workbook
from sklearn.preprocessing import LabelEncoder
from sklearn.utils.multiclass import type_of_target
import sys
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
from sklearn.model_selection import GridSearchCV
from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier

from real_datasets import load_dataset

from sklearn.gaussian_process import GaussianProcessRegressor, GaussianProcessClassifier
from sklearn.gaussian_process.kernels import RBF, ConstantKernel as C

import warnings
warnings.filterwarnings("ignore")

def train_gp(X, y, n_splits=5):
    """
    Train a Gaussian Process using k-fold cross-validation.
    
    Parameters:
        X (array-like): Feature matrix of shape (n_samples, n_features).
        y (array-like): Target vector of shape (n_samples,).
        n_splits (int): Number of folds for cross-validation (default is 5).
        
    Returns:
        avg_score (float): Average score (MSE or Accuracy) across folds.
        std_score (float): Standard deviation of the scores across folds.
    """
    
    # Check the type of problem (classification or regression) using sklearn utility
    problem_type = type_of_target(y)
    
    if problem_type == 'continuous' or problem_type == 'unknown':
        is_classification = False  # Regression problem
    else:
        is_classification = True   # Classification problem

    # Define the kernel for GP
    kernel = C(1.0, (1e-4, 1e1)) * RBF(1.0, (1e-4, 1e1))

    # Initialize the model
    if is_classification:
        gp = GaussianProcessClassifier(kernel=kernel, optimizer='fmin_l_bfgs_b')
    else:
        gp = GaussianProcessRegressor(kernel=kernel, optimizer='fmin_l_bfgs_b')

    # Setup K-fold Cross Validation
    kf = KFold(n_splits=n_splits, shuffle=True, random_state=42)

    # To store cross-validation results
    scores = []

    # Perform cross-validation
    for train_index, val_index in kf.split(X):
        X_train_fold, X_val_fold = X[train_index], X[val_index]
        y_train_fold, y_val_fold = y[train_index], y[val_index]
        
        # Train the GP on the training fold
        gp.fit(X_train_fold, y_train_fold)
        
        # Make predictions on the validation fold
        y_pred = gp.predict(X_val_fold)
        
        # Evaluate based on whether it's regression or classification
        if is_classification:
            # Calculate Accuracy for classification
            score = accuracy_score(y_val_fold, y_pred)
        else:
            # Calculate Mean Squared Error for regression
            score = mean_squared_error(y_val_fold, y_pred)
        
        scores.append(score)

    # Calculate average and standard deviation of scores
    avg_score = np.mean(scores)
    std_score = np.std(scores)

    # Print the results
    if is_classification:
        print(f"Average Accuracy across {n_splits}-folds: {avg_score}")
    else:
        print(f"Average MSE across {n_splits}-folds: {avg_score}")
        
    print(f"Standard Deviation of score: {std_score}")
    
    return avg_score, std_score

def reload_dataset(dataset_name):
    """Reload the dataset by its name."""
    return load_dataset(dataset_name)

def calculate_classification_scores(y_true, y_pred):
    """Calculate classification accuracy."""
    return {"Accuracy": accuracy_score(y_true, y_pred)}

def calculate_regression_scores(y_true, y_pred):
    """Calculate regression MSE."""
    return {"MSE": mean_squared_error(y_true, y_pred)}

def select_features_incrementally(X, y, ranked_features):
    """Select features incrementally and evaluate performance."""

    performance = []
    selected_features = []

    total_features = X.shape[1]
    i = math.ceil(total_features * 0.1)  # Start by selecting the top 10% of features

    while i <= total_features:
        # Select the top `i` ranked features
        selected_features = ranked_features[:i]

        # Subset the data for training and testing
        X_subset = X[:, selected_features]

        # Train the GP model on the selected features and evaluate
        avg_score, std_score = train_gp(X_subset, y)

        # Track performance for this number of selected features
        performance.append({
            'num_features': i,
            'avg_score': avg_score,
            'std_score': std_score
        })

        # Increment features by 10% of the total number of features
        i = math.ceil(total_features * 0.05 * (len(performance) + 1))  # Add 10% more each time

        # Ensure i does not exceed the total number of features
        if i > total_features:
            i = total_features

    return performance

def main():
    # Load the Excel file with feature importance data
    feature_importance_file = "feature_importance.xlsx"
    wb = load_workbook(feature_importance_file)

    # Create a new workbook for storing results
    results_wb = Workbook()

    # Check if an argument is passed
    ds_index = 1
    if len(sys.argv) < 2:
        print("No argument passed. Using default value: 1")
        ds_index = 1
    else:
        # Retrieve the parameter passed from Bash
        parameter = sys.argv[1]

        # Try converting the argument to a number
        try:
            # Try converting to an integer
            ds_index = int(parameter)
        except ValueError:
            # If it fails, try converting to a float
            ds_index = 1
            print("Cannot process the value. Using default value: 0.1")

    if ds_index == 1:
        sheetnames = wb.sheetnames[:5]
    elif ds_index == 2:
        sheetnames = wb.sheetnames[5:10]
    elif ds_index == 3:
        sheetnames = wb.sheetnames[10:15]


    # Process datasets in the Excel sheet
    for sheet_name in sheetnames:
        try:
            if sheet_name in ['keggdirected']: 
                continue 
            print(f"Processing dataset: {sheet_name}")
            sheet = wb[sheet_name]

            # Reload the dataset
            X, y = reload_dataset(sheet_name)

            # Determine if it's classification or regression
            is_classification = type_of_target(y) in ["binary", "multiclass"]
            if is_classification:
                y = LabelEncoder().fit_transform(y)

            result_sheet = results_wb.create_sheet(title=sheet_name)

            # Set column titles dynamically
            score_titles = ["Accuracy"] if is_classification else ["MSE"]
            result_sheet.append(["Feature Selector"] + score_titles)

            # Process each feature selector (row) in the sheet
            for row in sheet.iter_rows(min_row=2, values_only=True):
                feature_selector = row[0]
                if feature_selector is None:   break
                feature_importance = np.array(row[2:])

                # Use the ranking directly from the feature_importance array
                # Sorting by importance
                ranked_features = np.argsort(-np.abs(feature_importance))

                # Incrementally evaluate performance with the selected features
                performance = select_features_incrementally(X, y, ranked_features)

                # First row for the average scores
                avg_row = [feature_selector + "_avg"] + [result['avg_score'] for result in performance]
                result_sheet.append(avg_row)

                # Second row for the standard deviation scores
                std_row = [feature_selector + "_std"] + [result['std_score'] for result in performance]
                result_sheet.append(std_row)

            # Save the results to a new Excel file
            results_wb.save(f"class_svm_feature_selector_results_{ds_index}.xlsx")

        except Exception as e:
            print(f"{sheet_name} could not be processed! Error: {e}")
            continue


if __name__ == "__main__":
    main()
