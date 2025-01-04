import numpy as np
import matplotlib.pyplot as plt 

import shap 
from explainers.bishapley_kernel import Bivariate_KernelExplainer
from shapreg import removal, games, shapley
from explainers.MAPLE import MAPLE
from lime import lime_tabular
from pathlib import Path
import pandas as pd  
from openpyxl import load_workbook
from synthesized_data import *

from HSICNet.HSICFeatureNet import *
from HSICNet.HSICNet import *

from HSICNet.util import *

from explainers.L2x_reg import *
from invase import INVASE
from sklearn.linear_model import LinearRegression
import torch 

from explainers.L2x_reg import *
import os

# Define the output Excel file
output_file = "method_stats.xlsx"

# Check if the file exists to determine if it needs to be created or updated
if not os.path.exists(output_file):
    # Create an empty Excel file if it doesn't exist
    pd.DataFrame().to_excel(output_file, index=False)

# Create a wrapper class for 'fn' 
class CustomModel:
    def __init__(self, fn):
        self.fn = fn

    def fit(self, X, y=None):
        # Since this is a deterministic function, there's nothing to fit
        # This is just a placeholder to prevent errors in INVASE
        return self

    def predict(self, X):
        # Call the function 'fn' to get the target values 'y'
        return self.fn(X)


def create_rank(scores): 
	"""
	Compute rank of each feature based on weight.
	
	"""
	scores = abs(scores)
	n, d = scores.shape
	ranks = []
	for i, score in enumerate(scores):
		# Random permutation to avoid bias due to equal weights.
		idx = np.random.permutation(d) 
		permutated_weights = score[idx]  
		permutated_rank=(-permutated_weights).argsort().argsort()+1
		rank = permutated_rank[np.argsort(idx)]

		ranks.append(rank)

	return np.array(ranks)

# def performance_tp_fp(ranks, g_truth):

#     exists_features = np.zeros_like(ranks)
#     exists_features[np.argsort(ranks)[:np.sum(g_truth)]] = 1  # Predict top k ranked items as 1 (positive)

#     # True Positives (TP): Where both prediction and ground truth are 1
#     TP = np.sum((g_truth == 1) & (exists_features == 1))

#     # False Positives (FP): Where prediction is 1 but ground truth is 0
#     FP = np.sum((g_truth == 0) & (exists_features == 1))
#     return TP, FP


# def sort_shap_values(scores, k):
#     scores = np.abs(scores)
#     # Sort the array by absolute values in descending order and take the top two
#     top_k = scores[np.argsort(-scores)[:k]]
#     return 

def create_important_features_existence(ranks, g_truth):
    ''' ranks is the rank of each feature'''
    ''' This function finds the indices of the top k ranked features and 
        sets the corresponding positions in an important_features array to 1. '''

    important_features = np.zeros_like(ranks)
    for i in range(ranks.shape[0]):
        index_imp = np.argsort(ranks[i])[:int(np.sum(g_truth[i,:]))]
        important_features[i, index_imp] = 1 
    
    return important_features
    
def convert_Scores_to_impfExistence(score_init, Threshold):
    score_abs=abs(score_init)
    score = 1.*(score_abs > Threshold)    
    return score

import numpy as np

def compute_statistics(binary_array, ground_truth_indices):
    """
    Compute TP, FP, TN, FN for the entire dataset.

    Args:
        binary_array (np.ndarray): Array of shape (n_samples, n_variables) with binary values (0 or 1).
        ground_truth_indices (list or set): Indices of the variables that are true (selected in ground truth).

    Returns:
        dict: A dictionary containing TP, FP, TN, FN for the entire dataset.
    """
    n_variables = binary_array.shape[1]
    ground_truth_set = set(ground_truth_indices)
    
    # Initialize overall counts
    all_tp = len(ground_truth_set) * binary_array.shape[0]
    TP, FP, TN, FN = 0, 0, 0, 0
    
    for row in binary_array:
        selected_indices = set(np.where(row == 1)[0])  # Indices of selected variables
        
        # Update counts
        TP += len(selected_indices & ground_truth_set)  # True Positives
        FP += len(selected_indices - ground_truth_set)  # False Positives
        TN += len((set(range(n_variables)) - selected_indices) - ground_truth_set)  # True Negatives
        FN += len(ground_truth_set - selected_indices)  # False Negatives
    
    return TP, FP, TN, FN

def weight_to_binary(weight_values, n):
    """
    Converts weight values to a binary vector where only the n most important features
    by absolute weight value are 1, and the rest are 0.

    Parameters:
        shap_values (np.ndarray): Array of SHAP values for a single instance (1D array) or multiple instances (2D array).
        n (int): Number of most important features to select.

    Returns:
        np.ndarray: Binary vector(s) indicating the top n features (same shape as `shap_values`).
    """
    weight_values = np.atleast_2d(weight_values)  # Ensure 2D for consistency
    binary_vectors = np.zeros_like(weight_values, dtype=int)

    for i, instance_values in enumerate(weight_values):
        # Get the indices of the top-n absolute SHAP values
        top_indices = np.argsort(-np.abs(instance_values))[:n]
        binary_vectors[i, top_indices] = 1

    return binary_vectors if weight_values.shape[0] > 1 else binary_vectors[0]
    

def Compare_methods(X, y, X_test, X_sample_no, fn, feature_imp):
    n, d = X.shape
    input_dim = d

    epoch = 1

    X_tensor = torch.tensor(X, dtype=torch.float32)# torch.from_numpy(X).float()  # Convert to float tensor
    y_tensor = torch.tensor(y, dtype=torch.float32)  # Create float tensor directly from list or other data type
    X_tensor_test = torch.tensor(X_test, dtype=torch.float32)  # Create float tensor directly from list or other data type
    sigma_init_X = torch.tensor([0.5]*input_dim) #initialize_sigma_median_heuristic(X_tensor)
    sigma_init_Y = torch.tensor(0.5) #initialize_sigma_y_median_heuristic(y_tensor)
    num_samples = len(feature_imp)

    layers = [200, 300, 400, 500, 400, 200]
    feature_layers = [20, 50, 100, 50, 20]
    act_fun_featlayer = torch.nn.SELU
    act_fun_layer = torch.nn.Sigmoid
    
    featuregumbelsparsemax_model = HSICFeatureNetGumbelSparsemax(input_dim, feature_layers, act_fun_featlayer, layers, act_fun_layer, sigma_init_X, sigma_init_Y, num_samples * 3, temperature=20)
    featuregumbelsparsemax_model.train_model(X_tensor, y_tensor, num_epochs=epoch, BATCH_SIZE = 1000)
    weights = featuregumbelsparsemax_model(X_tensor)[0]
    HISCFGS_selected_features = (weights > 1e-3).to(torch.int32)
    HSICFNGS_stats  = compute_statistics(HISCFGS_selected_features, feature_imp)

    # HSICFeatureNetGumbelSparsemax
    featuregumbelsparsemax_model = HSICFeatureNetGumbelSparsemax(input_dim, feature_layers, act_fun_featlayer, layers, act_fun_layer, sigma_init_X, sigma_init_Y, num_samples * 5, temperature=20)
    featuregumbelsparsemax_model.train_model(X_tensor, y_tensor, num_epochs=epoch, BATCH_SIZE=1000)
    weights = featuregumbelsparsemax_model(X_tensor)[0]
    HISCFGS_selected_features = (weights > 1e-3).to(torch.int32)
    HSICFNGS2_stats = compute_statistics(HISCFGS_selected_features, feature_imp)

    # HSICNetGumbelSparsemax
    gumbelsparsemax_model = HSICNetGumbelSparsemax(input_dim, layers, act_fun_layer, sigma_init_X, sigma_init_Y, num_samples)
    gumbelsparsemax_model.train_model(X_tensor, y_tensor, num_epochs=epoch, BATCH_SIZE=1000)
    weights = gumbelsparsemax_model(X_tensor)[0]
    gumbelsparsemax_selected_features = (weights > 1e-3).to(torch.int32)
    HSIC_gumbsparsemax_stats = compute_statistics(gumbelsparsemax_selected_features, feature_imp)
    
    # HSICNetGumbelSparsemax with different sampling
    gumbelsparsemax_model2 = HSICNetGumbelSparsemax(input_dim, layers, act_fun_layer, sigma_init_X, sigma_init_Y, num_samples*2)
    gumbelsparsemax_model2.train_model(X_tensor, y_tensor, num_epochs=epoch, BATCH_SIZE=1000)
    weights2 = gumbelsparsemax_model2(X_tensor)[0]
    gumbelsparsemax2_selected_features = (weights2 > 1e-3).to(torch.int32)
    HSIC_gumbsparsemax2_stats = compute_statistics(gumbelsparsemax2_selected_features, feature_imp)
    
    # HSICNetGumbelSoftmax
    gumbelsoftmax_model = HSICNetGumbelSoftmax(input_dim, layers, act_fun_layer, sigma_init_X, sigma_init_Y, num_samples)
    gumbelsoftmax_model.train_model(X_tensor, y_tensor, num_epochs=epoch, BATCH_SIZE=1000)
    weights = gumbelsoftmax_model.importance_weights
    gumbelsoftmax_selected_features = (weights > 1e-3).to(torch.int32)
    HSIC_gumbsoftmax_stats = compute_statistics(gumbelsoftmax_selected_features, feature_imp)
    
    # HSICNetGumbelSoftmax with 2*num_samples
    gumbelsoftmax_model2 = HSICNetGumbelSoftmax(input_dim, layers, act_fun_layer, sigma_init_X, sigma_init_Y, 2 * num_samples)
    gumbelsoftmax_model2.train_model(X_tensor, y_tensor, num_epochs=epoch, BATCH_SIZE=1000)
    weights2 = gumbelsoftmax_model2.importance_weights
    gumbelsoftmax2_selected_features = (weights2 > 1e-3).to(torch.int32)
    HSIC_gumbsoftmax2_stats = compute_statistics(gumbelsoftmax2_selected_features, feature_imp)
    
    # HSICNetSparsemax
    sparsemax_model = HSICNetSparsemax(input_dim, layers, act_fun_layer, sigma_init_X, sigma_init_Y)
    sparsemax_model.train_model(X_tensor, y_tensor, num_epochs=epoch, BATCH_SIZE=1000)
    weights = sparsemax_model.importance_weights
    sparsemax_selected_features = (weights > 1e-3).to(torch.int32)
    HSIC_sparsemax_stats = compute_statistics(sparsemax_selected_features, feature_imp)

    # Training INVASE
    feature_names = [f"Feature_{i}" for i in range(X.shape[1])]
    X_df = pd.DataFrame(X, columns=feature_names)
    y_series = pd.Series(y, name="Target")
    model = CustomModel(fn)
    Invase_explainer = INVASE (model, X_df, y_series, n_epoch=epoch, prefit=False)
    invase_scores =(Invase_explainer.explain(X_df)).to_numpy()                      
    invase_selected_features = (invase_scores > 0.5).astype(int)
    invase_stats  = compute_statistics(invase_selected_features, feature_imp)

    #L2X #retrun feature importance
    l2x_weights = train_L2X(X_tensor, y_tensor, len(feature_imp), epochs=epoch, batch_size=BATCH_SIZE)
    L2X_selected_features = (l2x_weights > 1e-3).astype(int)
    L2x_stats = compute_statistics(L2X_selected_features, feature_imp)
    
    #L2X with 2*num_feature_imp
    l2x_2_weights = train_L2X(X_tensor, y_tensor, 2*len(feature_imp), epochs=epoch, batch_size=BATCH_SIZE)
    L2X_2_selected_features = (l2x_2_weights > 1e-3).astype(int)
    l2x_2_stats = compute_statistics(L2X_2_selected_features, feature_imp)
    
    ## SHAP
    X_tbx = X[:20,:]
    X_bg = shap.sample(X, 100)
    explainer = shap.KernelExplainer(fn, X_bg)
    shap_values = explainer.shap_values(X_tbx, nsamples=X_sample_no, l1_reg=True)
    shap_selected_featuers = (np.abs(shap_values) > 1e-3).astype(int)
    shap_stats = compute_statistics(shap_selected_featuers, feature_imp)
    
    ## Bivariate SHAP
    bishap = Bivariate_KernelExplainer(fn, X_bg)
    bishap_values = bishap.shap_values(X_tbx, nsamples=X_sample_no, l1_reg=True)
    bishap_selected_featuers = (np.abs(bishap_values) > 1e-3).astype(int)
    bishap_stats = compute_statistics(bishap_selected_featuers, feature_imp)

    ## LIME, Unbiased SHAP, and MAPLE 
    lime_exp = lime_tabular.LimeTabularExplainer(X_bg, discretize_continuous=False, mode="regression")

    lime_values = np.empty_like(X_tbx)
    for i in range(X_tbx.shape[0]):
        x = X_tbx[i, ]
        ## LIME 
        exp = lime_exp.explain_instance(x, fn, num_samples = X_sample_no)
            
        for tpl in exp.as_list():
            lime_values[i, int(tpl[0])] = tpl[1]

    lime_selected_featuers = (np.abs(lime_values) > 1e-3).astype(int)
    lime_stats = compute_statistics(lime_selected_featuers, feature_imp)

    return HSICFNGS_stats, HSICFNGS2_stats, HSIC_gumbsparsemax_stats, HSIC_gumbsparsemax2_stats, \
           HSIC_gumbsoftmax_stats, HSIC_gumbsoftmax2_stats, HSIC_sparsemax_stats, invase_stats, \
           L2x_stats, l2x_2_stats, shap_stats, bishap_stats, lime_stats

if __name__=='__main__':

    
    num_samples = 2000 # number of generated synthesized instances 
    input_dim = 20 # number of features for the synthesized instances
    hidden_dim1 = 100
    hidden_dim2 = 100
    X_sample_no = 200  # number of sampels for generating explanation
    train_seed = 42
    test_seed = 1
    Threshold = 0.01

   
    # data_sets=['Sine Log', 'Sine Cosine', 'Poly Sine', 'Squared Exponentials', 'Tanh Sine', 
    #          'Trigonometric Exponential', 'Exponential Hyperbolic', 'XOR', 'Syn4']
    # ds_name = data_sets[1]
    # data_sets= ['Syn4']
    data_sets=['Squared Exponentials']

    for ds_name in data_sets:
        ## Not good: Trigonometric Exponential (INVASE not good either), XOR,
        ## Good: Poly Sine, Squared Exponentials, Sine Log,Sine Cosine, Exponential Hyperbolic
        # Generate synthetic data
        X_train, y_train, fn, feature_imp, g_train = generate_dataset(ds_name, num_samples, input_dim, train_seed)
        # X_test, y_test, fn, feature_imp, g_test = generate_dataset(ds_name, num_samples, input_dim, test_seed)
        
        #HSICFNGS_stats, HSICFNGS2_stats, HSIC_gumbsparsemax_stats, HSIC_gumbsparsemax2_stats, \
        #HSIC_gumbsoftmax_stats, HSIC_gumbsoftmax2_stats, HSIC_sparsemax_stats, invase_stats, \
        #L2x_stats, l2x_2_stats, shap_stats, bishap_stats, lime_stats = Compare_methods(X_train, y_train, X_train, X_sample_no, fn, feature_imp)
        
        stats = Compare_methods(X_train, y_train, X_train, X_sample_no, fn, feature_imp)
    
    # Methods names corresponding to the stats
        methods = [
            "HSIC_FN_GS", "HSIC_FN_GS2", "HSIC_gumbsparsemax", "HSIC_gumbsparsemax2",
            "HSIC_gumbsoftmax", "HSIC_gumbsoftmax2", "HSIC_sparsemax", "invase",
            "L2x", "l2x_2", "shap", "bishap", "lime"
        ]
        
        # Prepare data for the DataFrame
        data = []
        for method, stat in zip(methods, stats):
            data.append({
                "Method": method,
                "TP": stat[0],
                "FP": stat[1],
                "TN": stat[2],
                "FN": stat[3],
            })
        
        # Create a DataFrame
        df = pd.DataFrame(data)
        
        # Load the existing Excel file
        book = load_workbook(output_file)
        
        # Remove the sheet if it already exists
        if ds_name in book.sheetnames:
            del book[ds_name]
        
        # Write the DataFrame to a new sheet
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, index=False, sheet_name=ds_name)


        print("done!")


