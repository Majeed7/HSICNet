import numpy as np
import pandas as pd
from sklearn.feature_selection import mutual_info_classif, RFECV, SelectKBest, mutual_info_regression
from sklearn.linear_model import Lasso
from sklearn.svm import SVC, SVR
from sklearn.ensemble import ExtraTreesClassifier, ExtraTreesRegressor
from sklearn.feature_selection import SelectKBest, f_classif, f_regression
#from pyHSICLasso import HSICLasso
from pathlib import Path
import os 
from openpyxl import load_workbook
import time

from HSICNet.HSICFeatureNet import *
from HSICNet.HSICNet import *

from synthesized_data import *

#import warnings
#warnings.filterwarnings("ignore")

import torch
print("CUDA available:", torch.cuda.is_available())
print("CUDA version:", torch.version.cuda)

results_xsl = Path('hsic_fs_synthesized.xlsx')
if not os.path.exists(results_xsl):
    # Create an empty Excel file if it doesn't exist
    pd.DataFrame().to_excel(results_xsl, index=False)

'''
if __name__ == '__main__':
    np.random.seed(30)

    sample_no_gn = 1000 # number of generated synthesized instances 
    feature_no_gn = 16 # number of features for the synthesized instances

    exp_no = 30 # number of experiments
    importance_mi = np.zeros((exp_no,feature_no_gn))
    importance_lasso = np.zeros((exp_no,feature_no_gn))
    orders_rfecv = np.zeros((exp_no,feature_no_gn))
    importance_k_best = np.zeros((exp_no,feature_no_gn))
    importance_ensemble = np.zeros((exp_no,feature_no_gn))
    importance_hsiclasso = np.zeros((exp_no,feature_no_gn))
    importance_hsicfngs = np.zeros((exp_no,feature_no_gn))
    importance_hsicfngs2 = np.zeros((exp_no,feature_no_gn))
    importance_hsicgs = np.zeros((exp_no,feature_no_gn))
    importance_hsicgs2 = np.zeros((exp_no,feature_no_gn))

    time_mi = np.zeros(exp_no)
    time_lasso = np.zeros(exp_no)
    time_rfecv = np.zeros(exp_no)
    time_k_best = np.zeros(exp_no)
    time_ensemble = np.zeros(exp_no)
    time_hsiclasso = np.zeros(exp_no)
    time_hsicfngs = np.zeros(exp_no)
    time_hsicfngs2 = np.zeros(exp_no)
    time_hsicgs = np.zeros(exp_no)
    time_hsicgs2 = np.zeros(exp_no)

    # Example usage of one of the functions
    datasets=['Sine Log', 'Sine Cosine', 'Poly Sine', 'Squared Exponentials', 'Tanh Sine', 
              'Trigonometric Exponential', 'Exponential Hyperbolic', 'XOR'] #
    
    epoch=500
    layers = [200, 300, 200]
    feature_layers = [20, 50, 20]
    act_fun_featlayer = torch.nn.SELU
    act_fun_layer = torch.nn.Sigmoid

    for ds_name in datasets:
        for i in range(exp_no):
            X, y, fn, feature_imp, g_train = generate_dataset(ds_name, sample_no_gn, feature_no_gn, 42)

            mode = 'regression'
            print(ds_name, i)
            ## HSIC based feature selection
            X_tensor = torch.tensor(X, dtype=torch.float32)# torch.from_numpy(X).float()  # Convert to float tensor
            y_tensor = torch.tensor(y, dtype=torch.float32)  # Create float tensor directly from list or other data type
            sigma_init_X = torch.tensor([0.5]*feature_no_gn) #initialize_sigma_median_heuristic(X_tensor)
            sigma_init_Y = torch.tensor(0.5) #initialize_sigma_y_median_heuristic(y_tensor)
            num_samples = len(feature_imp)

            start_time = time.time()
            featuregumbelsparsemax_model = HSICFeatureNetGumbelSparsemax(feature_no_gn, feature_layers, act_fun_featlayer, layers, act_fun_layer, sigma_init_X, sigma_init_Y, num_samples * 3, temperature=20)
            featuregumbelsparsemax_model.train_model(X_tensor, y_tensor, num_epochs=epoch, BATCH_SIZE = 1000)
            weights = featuregumbelsparsemax_model(X_tensor)[0]
            hsicfngs_sv, v0 = featuregumbelsparsemax_model.global_shapley_value(X_tensor, y_tensor, featuregumbelsparsemax_model.sigmas, featuregumbelsparsemax_model.sigma_y, weights)
            importance_hsicfngs[i,:] = hsicfngs_sv.detach().numpy().squeeze()
            time_hsicfngs[i] = time.time() - start_time
            
                ## HSICNetGumbelSparsemax
            gumbelsparsemax_model = HSICNetGumbelSparsemax(feature_no_gn, layers, act_fun_layer, sigma_init_X, sigma_init_Y, num_samples)
            gumbelsparsemax_model.train_model(X_tensor, y_tensor, num_epochs=epoch, BATCH_SIZE=1000)
            weights = gumbelsparsemax_model(X_tensor)[0]
            hsicgs_sv, v0 = featuregumbelsparsemax_model.global_shapley_value(X_tensor, y_tensor, featuregumbelsparsemax_model.sigmas, featuregumbelsparsemax_model.sigma_y, weights)
            importance_hsicgs[i,:] = hsicgs_sv.detach().numpy().squeeze()
            time_hsicgs[i] = time.time() - start_time
            
            ## HSIC lasso 
            # start_time = time.time()
            # hsic_lasso = HSICLasso()
            # hsic_lasso.input(X,y)
            # hsic_lasso.regression(feature_no_gn, covars=X)
            # hsic_ind = hsic_lasso.get_index()
            # init_ranks = (len(hsic_ind) + (feature_no_gn - 1/2 - len(hsic_ind))/2) * np.ones((feature_no_gn,))
            # init_ranks[hsic_ind] = np.arange(1,len(hsic_ind)+1)
            # importance_hsiclasso[i,:] = hsicgs_sv.detach().numpy().squeeze()  
            # time_hsiclasso[i] = time.time() - start_time 

            ## Mutual Informmation Importance
            start_time = time.time()
            importance_mi[i,:] = mutual_info_classif(X,y) if mode == 'classification' else mutual_info_regression(X,y)
            time_mi[i] = time.time() - start_time

            ## Lasso importance
            start_time = time.time()
            lasso = Lasso().fit(X, y)
            importance_lasso[i,:] = np.abs(lasso.coef_)
            time_lasso[i] = time.time() - start_time

            #Recursive elimination
            start_time = time.time()
            estimator = SVC(kernel="linear") if mode == 'classification' else SVR(kernel='linear')
            rfecv = RFECV(estimator, step=1, cv=5)
            rfecv.fit(X, y)
            orders_rfecv[i,:] = rfecv.ranking_
            time_rfecv[i] = time.time() - start_time    

            ## K best
            start_time = time.time()
            bestfeatures = SelectKBest(score_func=f_classif, k='all') if mode == 'classification' else SelectKBest(score_func=f_regression, k='all') #F-ANOVA feature selection
            fit = bestfeatures.fit(X,y)
            importance_k_best[i,:] = fit.scores_
            time_k_best[i] = time.time() - start_time   

            ## Tree ensemble 
            start_time = time.time()
            model = ExtraTreesClassifier(n_estimators=50) if mode =='classification' else ExtraTreesRegressor(n_estimators=50)
            model.fit(X,y)
            importance_ensemble[i,:] = model.feature_importances_
            time_ensemble[i] = time.time() - start_time 

        ranking_mi = create_rank(importance_mi)
        ranking_lasso = create_rank(importance_lasso)
        ranking_k_best = create_rank(importance_k_best)
        ranking_ensemble = create_rank(importance_ensemble)
        ranking_rfecv = create_rank(orders_rfecv)
        ranking_hsiclasso = importance_hsiclasso
        ranking_hsicfngs = create_rank(np.abs(importance_hsicfngs))
        ranking_hsicfngs2 = create_rank(np.abs(importance_hsicfngs2))
        ranking_hsicgs = create_rank(np.abs(importance_hsicgs))
        ranking_hsicgs2 = create_rank(np.abs(importance_hsicgs2))
        
        avg_mi = (np.mean(ranking_mi[:,feature_imp],axis=1))
        avg_lasso = (np.mean(ranking_lasso[:,feature_imp],axis=1))
        avg_k_best = (np.mean(ranking_k_best[:,feature_imp],axis=1))
        avg_ensemble = (np.mean(ranking_ensemble[:,feature_imp],axis=1))
        avg_rfecv = (np.mean(ranking_rfecv[:,feature_imp],axis=1))
        avg_hsic_lasso = (np.mean(ranking_hsiclasso[:,feature_imp],axis=1))
        avg_hsicfngs = (np.mean(ranking_hsicfngs[:,feature_imp],axis=1))
        avg_hsicfngs2 = (np.mean(ranking_hsicfngs2[:,feature_imp],axis=1))
        avg_hsicgs = (np.mean(ranking_hsicgs[:,feature_imp],axis=1))
        avg_hsicgs2 = (np.mean(ranking_hsicgs2[:,feature_imp],axis=1))  
    
        # Creating dataset
        data = [avg_hsic_lasso, time_hsiclasso, avg_mi, time_mi, avg_k_best, time_k_best, avg_rfecv, time_rfecv, avg_lasso, time_lasso, avg_ensemble, time_ensemble, \
                avg_hsicfngs, time_hsicfngs, avg_hsicfngs2, time_hsicfngs2, avg_hsicgs, time_hsicgs, avg_hsicgs2, time_hsicgs2]
        methods = ["HSIC-Lasso", "HSIC_time", "MI", "MI_time", "F-ANOVA", "F_ANOVA_time", "RFEC", "RFEC_time", "Lasso", "Lasso_time", "Tree Ensemble", "Tree_ensemble_time", \
                    "HSICFNGS", "HSICFNGS_time", "HSICFNGS2", "HSICFNGS2_time", "HSICGS", "HSICGS_time", "HSICGS2", "HSICGS2_time"]

        df = pd.DataFrame(data, index=methods)

        mode = 'a' if results_xsl.exists() else 'w'
        # Load the existing Excel file
        book = load_workbook(results_xsl)
        
        # Remove the sheet if it already exists
        if ds_name in book.sheetnames:
            del book[ds_name]
        
        # Write the DataFrame to a new sheet
        with pd.ExcelWriter(results_xsl, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, index=False, sheet_name=ds_name)

'''
